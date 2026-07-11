#!/usr/bin/env python3
"""Script to log daily scrum standup notes to an Excel spreadsheet."""

import os
import sys
import datetime
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Installing openpyxl dependency...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKSPACE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(WORKSPACE_DIR, "docs/daily_standups.xlsx")

STANDUP_DATA = [
    {
        "Role": "Business Analyst",
        "Work Completed": "Formulated requirements and drafted initial solution brief (5 target classes, YOLOv8 vehicle filtering, crop padding, 0.75 threshold).",
        "Next Steps": "Prepare user stories for Phase 5 integration testing.",
        "Blockers": "None"
    },
    {
        "Role": "UX Designer",
        "Work Completed": "Designed layout wireframes for Dashboard and Live Camera view. Produced high-fidelity Mock Screen specifications (Obsidian Dark palette, glassmorphism CSS styling, hover interactions).",
        "Next Steps": "Review UI implementations to ensure design fidelity.",
        "Blockers": "None"
    },
    {
        "Role": "Solution Architect",
        "Work Completed": "Reviewed and approved UX wireframe layout for technical execution. Validated canvas overlays and 10-15 FPS frame throttling constraints.",
        "Next Steps": "Support integration of Django endpoints with Angular UI.",
        "Blockers": "None"
    },
    {
        "Role": "ML Engineer",
        "Work Completed": "Audited dataset (Audi: 92, BMW: 84, Jeep: 85, Suzuki: 80, Tesla: 77). Configured ResNet50 classifier training scripts and completed all 50 Epochs. Model achieved 93.75% test accuracy.",
        "Next Steps": "Idle / Support QA and integration verification.",
        "Blockers": "None"
    },
    {
        "Role": "Backend Developer",
        "Work Completed": "Initialized Django project, configured settings.py/urls.py, installed dependencies. Created views.py with YOLOv8 detection and ResNet50 hooks. Completed model integration with verified best ResNet50 model weights.",
        "Next Steps": "Support QA end-to-end integration testing.",
        "Blockers": "None"
    },
    {
        "Role": "Frontend Developer",
        "Work Completed": "Initialized Angular workspace, installed npm packages. Created CarDetectionService, router routes, and implemented layouts/templates for Dashboard and LiveCamera components.",
        "Next Steps": "Support QA end-to-end integration testing.",
        "Blockers": "None"
    },
    {
        "Role": "Scrum Master",
        "Work Completed": "Facilitated Scrum Call Day 1, logged team progress, and created standup tracking files (markdown and Excel logs). Verified successful end-to-end integration API testing.",
        "Next Steps": "Support final CEO review and project release gates.",
        "Blockers": "None"
    }
]

def main():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    # Load or create workbook
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Standup Log"
        # Write headers
        headers = ["Date", "Sprint", "Role", "Work Completed", "Next Steps", "Blockers"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num in range(1, 7):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        ws.row_dimensions[1].height = 28

    today_str = datetime.date.today().strftime("%Y-%m-%d")
    sprint_str = "Sprint 1"
    
    # Styles for data rows
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    border_side = Side(style='thin', color='D3D3D3')
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    start_row = ws.max_row + 1
    
    for entry in STANDUP_DATA:
        row_data = [
            today_str,
            sprint_str,
            entry["Role"],
            entry["Work Completed"],
            entry["Next Steps"],
            entry["Blockers"]
        ]
        ws.append(row_data)
        
        # Style the new row
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 45
        for col_num in range(1, 7):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border
            
            # Align first 3 columns center
            if col_num in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 12

    wb.save(EXCEL_PATH)
    print(f"Logged daily standup data to Excel: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
