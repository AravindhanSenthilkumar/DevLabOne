#!/usr/bin/env python3
"""Script to log story tracking assignments to an Excel spreadsheet."""

import os
import sys
import datetime
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Installing openpyxl dependency...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKSPACE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(WORKSPACE_DIR, "docs/story_tracker.xlsx")

STORIES_DATA = [
    {
        "ID": "ST-01",
        "Role": "Business Analyst",
        "Story Title": "Solution Brief Formulation",
        "Description": "Formulate requirements and write the solution proposal brief for YOLOv8 detection & ResNet50 classification.",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-02",
        "Role": "UX Designer",
        "Story Title": "UX Wireframe Layouts Design",
        "Description": "Create layout wireframes for the Upload Dashboard and Live Monitoring views. [Linked Diagram: docs/architecture_flow.svg]",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-03",
        "Role": "UX Designer",
        "Story Title": "High-Fidelity Mock Screens Design",
        "Description": "Define design tokens, HSL colors, Obsidian Dark styling guidelines, and card hover interactions. [Linked Mockup: docs/ux_ui_design.svg]",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-04",
        "Role": "Solution Architect",
        "Story Title": "Wireframe Feasibility Audit",
        "Description": "Review UX wireframes for coordinates mapping scalability and live frame rate throttling bounds. [Linked Flow: docs/architecture_wireframe.md]",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-05",
        "Role": "ML Engineer",
        "Story Title": "ResNet50 Classifier Training",
        "Description": "Audit dataset splits, setup torch configurations, and run training comparison for 10/20/30/40/50 epochs.",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-06",
        "Role": "Backend Developer",
        "Story Title": "Django API & Views Setup",
        "Description": "Initialize django backend, write REST endpoints, and implement YOLOv8 detection/crop views.",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-07",
        "Role": "Frontend Developer",
        "Story Title": "Angular UI Components Coding",
        "Description": "Initialize Angular app, build CarDetectionService, Dashboard upload views, and LiveCamera canvas drawing stream.",
        "Status": "DONE",
        "Blocker Note": "-"
    },
    {
        "ID": "ST-08",
        "Role": "Scrum Master",
        "Story Title": "Standup Logs & Trackers Setup",
        "Description": "Establish Excel and Markdown trackers for Standup logs and daily story assignments.",
        "Status": "DONE",
        "Blocker Note": "-"
    }
]

def main():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    # Load or create workbook
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Story Tracker"
        # Write headers
        headers = ["Date", "Story ID", "Role", "Story Title", "Description", "Status", "Blocker Note"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num in range(1, 8):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        ws.row_dimensions[1].height = 28

    today_str = datetime.date.today().strftime("%Y-%m-%d")
    
    # Styles for data rows
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    border_side = Side(style='thin', color='D3D3D3')
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Colors for statuses
    status_colors = {
        "DONE": "D1FAE5",         # Light emerald
        "READY FOR REVIEW": "DBEAFE", # Light blue
        "IN PROGRESS": "FEF3C7",   # Light amber
        "TO DO": "F3F4F6",         # Light gray
        "BLOCKER": "FEE2E2"        # Light red
    }
    
    for entry in STORIES_DATA:
        row_data = [
            today_str,
            entry["ID"],
            entry["Role"],
            entry["Story Title"],
            entry["Description"],
            entry["Status"],
            entry["Blocker Note"]
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 40
        for col_num in range(1, 8):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border
            
            # Center alignments
            if col_num in [1, 2, 3, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Highlight status cell
            if col_num == 6:
                status = entry["Status"]
                color_hex = status_colors.get(status, "FFFFFF")
                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell.font = Font(name="Arial", size=10, bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 35

    wb.save(EXCEL_PATH)
    print(f"Logged story tracker data to Excel: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
